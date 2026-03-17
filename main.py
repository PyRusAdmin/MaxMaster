# -*- coding: utf-8 -*-
"""
MaxMaster - приложение для парсинга пользователей мессенджера Max по номерам телефона.

Основные возможности:
- Подключение аккаунтов Max через QR-код
- Перебор номеров телефона из очереди
- Поиск пользователей по номеру телефона
- Сохранение результатов в Excel
- Автоматическая ротация аккаунтов при блокировках
- Обработка rate limit ограничений

Автор: MaxMaster Team
Версия: 0.0.1
"""
import asyncio
import os
import sys
from datetime import datetime

from loguru import logger
from openpyxl import Workbook
from rich import box
from rich.align import Align
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.prompt import Prompt
from rich.table import Table
from rich.text import Text

from PyMax.src.pymax.core import MaxClient
from PyMax.src.pymax.payloads import UserAgentPayload
from config import DB_PATH, EXCEL_FILE, APP_VERSION, APP_DATE, SLEEP_TIME, SLEEP_ON_RATELIMIT, NUMBERS_FILE
from database import AccountLog, MaxAccount, PhoneQueue
from read_file import read_file

# ─── Логгер ───────────────────────────────────────────────────────────────────
# Настройка логгера: запись в файл и вывод в консоль
logger.remove()  # Удаляем стандартный обработчик
logger.add("logs/log.log", rotation="1 MB", level="INFO")  # Запись в файл с ротацией по размеру
logger.add(sys.stderr, level="WARNING")  # Вывод предупреждений и ошибок в консоль

# ─── Rich консоль ─────────────────────────────────────────────────────────────
# Инициализация консоли для красивого вывода с поддержкой цветов и форматирования
console = Console()

# ─── Max клиент ───────────────────────────────────────────────────────────────
# Глобальное хранилище клиентов: {phone: MaxClient}
# Кэширует подключённых клиентов для повторного использования
active_clients: dict[str, MaxClient] = {}


async def client_connect(phone: str | None = None, work_dir: str = "accounts", timeout: float = 30.0) -> MaxClient:
    """
    Создаёт и возвращает клиента MaxClient.

    Если клиент для этого телефона уже существует — возвращает его.

    :param phone: Номер телефона аккаунта Max.
    :param work_dir: Рабочая директория для хранения базы данных с аккаунтами.
    :param timeout: Таймаут подключения в секундах.
    :return: Экземпляр MaxClient.
    :raises TimeoutError: Если подключение не удалось в течение timeout.
    """
    # Проверяем, есть ли уже клиент для этого телефона (кэширование)
    if phone and phone in active_clients:
        client = active_clients[phone]
        if client.is_connected:
            logger.info(f"Используем существующего клиента для {phone}")
            return client
        else:
            # Клиент отключён, удаляем из кэша
            del active_clients[phone]

    # Настройка заголовков User-Agent для веб-устройства
    headers = UserAgentPayload(device_type="WEB")

    # Создание клиента Max
    client = MaxClient(
        phone=phone,  # Номер телефона аккаунта
        work_dir=work_dir,  # Рабочая директория для хранения базы данных с аккаунтами Max
        reconnect=True,  # Включаем авто-переподключение при разрыве соединения
        headers=headers,
    )

    # Запускаем клиента
    console.print(f"[dim]⏳ Подключение к {phone} (таймаут {timeout}с)...[/]")

    # Запускаем start() в фоне и ждём подключения
    start_task = asyncio.create_task(client.start())

    # Ждём подключения с таймаутом
    try:
        for i in range(int(timeout * 10)):  # Проверяем каждые 100мс
            await asyncio.sleep(0.1)

            if client.is_connected:
                logger.info(f"✅ Клиент {phone} подключён")
                break

            if start_task.done() and start_task.exception():
                exc = start_task.exception()
                logger.error(f"Ошибка при старте клиента {phone}: {exc}")
                raise exc

        # Проверяем, подключился ли клиент
        if not client.is_connected:
            raise TimeoutError(f"Не удалось подключиться к {phone} за {timeout}с")

    except asyncio.CancelledError:
        logger.error(f"Подключение {phone} отменено")
        raise
    except Exception as e:
        logger.error(f"Ошибка подключения для {phone}: {e}")
        console.print(f"[red]❌ Ошибка подключения: {e}[/]")
        # Отменяем задачу старта
        start_task.cancel()
        try:
            await client.close()
        except Exception:
            pass
        raise

    # Сохраняем в глобальное хранилище (кэш)
    if phone:
        active_clients[phone] = client
        logger.info(f"✅ Клиент для {phone} подключён и сохранён в кэш")

    return client


async def client_disconnect(phone: str) -> bool:
    """
    Отключает клиента и удаляет из кэша.

    :param phone: Номер телефона аккаунта.
    :return: True если успешно отключён.
    """
    if phone in active_clients:
        client = active_clients[phone]
        try:
            await client.close()  # Закрываем соединение с сервером Max
            del active_clients[phone]  # Удаляем из кэша
            logger.info(f"Клиент {phone} отключён")
            return True
        except Exception as e:
            logger.error(f"Ошибка отключения клиента {phone}: {e}")
    return False


async def disconnect_all_clients():
    """Отключает всех клиентов из глобального хранилища."""
    for phone in list(active_clients.keys()):
        await client_disconnect(phone)


# ─── Управление аккаунтами ────────────────────────────────────────────────────

def log_account_action(phone: str, action: str, message: str = None):
    """
    Логирует действие с аккаунтом в базу данных.
    
    :param phone: Номер телефона аккаунта.
    :param action: Тип действия (start/work/error/block/stop).
    :param message: Дополнительное сообщение.
    """
    AccountLog.create(phone=phone, action=action, message=message)


def get_active_accounts() -> list:
    """
    Получает список активных незаблокированных аккаунтов из базы данных.
    
    :return: Список объектов MaxAccount, отсортированных по количеству ошибок и времени использования.
    """
    return list(MaxAccount.select().where(
        (MaxAccount.is_active == "Y") &  # Только активные
        (MaxAccount.is_blocked == "N") &  # Только незаблокированные
        (MaxAccount.work_status != "blocked")  # Исключая заблокированные по статусу
    ).order_by(MaxAccount.errors_count.asc(), MaxAccount.last_used_at.asc()))


def mark_account_working(phone: str):
    """
    Отмечает аккаунт как работающий (в процессе работы).
    
    :param phone: Номер телефона аккаунта.
    """
    MaxAccount.update(
        work_status="working",  # Статус: в работе
        last_used_at=datetime.now()  # Обновляем время последнего использования
    ).where(MaxAccount.phone == phone).execute()
    log_account_action(phone, "start", "Начал работу")


def mark_account_idle(phone: str):
    """
    Отмечает аккаунт как свободный (завершил работу).
    
    :param phone: Номер телефона аккаунта.
    """
    MaxAccount.update(work_status="idle").where(MaxAccount.phone == phone).execute()
    log_account_action(phone, "stop", "Завершил работу")


def mark_account_error(phone: str):
    """
    Регистрирует ошибку аккаунта и увеличивает счётчик ошибок.
    Если ошибок больше 3 — блокирует аккаунт.
    
    :param phone: Номер телефона аккаунта.
    :return: True если аккаунт заблокирован, иначе False.
    """
    account = MaxAccount.get_or_none(MaxAccount.phone == phone)
    if account:
        errors = int(account.errors_count) + 1
        MaxAccount.update(errors_count=str(errors)).where(MaxAccount.phone == phone).execute()
        log_account_action(phone, "error", f"Ошибка №{errors}")

        # Если больше 3 ошибок — блокируем аккаунт
        if errors >= 3:
            MaxAccount.update(
                work_status="blocked",
                is_blocked="Y"
            ).where(MaxAccount.phone == phone).execute()
            log_account_action(phone, "block", f"Заблокирован после {errors} ошибок")
            return True  # Аккаунт заблокирован
    return False


def mark_account_blocked(phone: str, reason: str = "Блокировка сервера"):
    """
    Блокирует аккаунт по указанной причине.
    
    :param phone: Номер телефона аккаунта.
    :param reason: Причина блокировки.
    """
    MaxAccount.update(
        work_status="blocked",
        is_blocked="Y"
    ).where(MaxAccount.phone == phone).execute()
    log_account_action(phone, "block", reason)


# ─── Подключение аккаунта по QR ───────────────────────────────────────────────

async def connect_account_by_qr() -> bool:
    """
    Подключает аккаунт Max через QR-код.

    Процесс подключения:
    1. Пользователь вводит номер телефона
    2. Создаётся папка для хранения данных аккаунта
    3. Генерируется QR-код для сканирования в приложении Max
    4. После сканирования аккаунт сохраняется в базу данных

    :return: True если подключение успешно, иначе False.
    """
    # Вывод инструкции по подключению для пользователя
    console.print(Panel(
        "[bold cyan]🔌 Подключение аккаунта Max[/]\n\n"
        "Для подключения отсканируйте QR-код в приложении Max.\n\n"
        "[yellow]Важно:[/]\n"
        "  • QR-код действителен 60 секунд\n"
        "  • Откройте приложение Max\n"
        "  • Перейдите в настройки\n"
        "  • Выберите 'Подключить устройство'\n",
        title="[bold green]Инструкция[/]",
        border_style="green",
        padding=(1, 3),
    ))

    # Запрос номера телефона у пользователя
    phone = Prompt.ask(
        "[bold yellow]Введите номер телефона (например, 998950039094)[/]",
        default="",
    )

    if not phone:
        console.print("[red]❌ Номер телефона не введён[/]")
        return False

    # Очищаем номер от лишних символов (оставляем только цифры)
    phone = ''.join(filter(str.isdigit, phone))

    # Создаём папку для хранения данных аккаунта
    account_path = os.path.join("accounts", phone)
    os.makedirs(account_path, exist_ok=True)

    try:
        # Настройка заголовков User-Agent для веб-устройства
        headers = UserAgentPayload(device_type="WEB")

        # Создание клиента Max
        client = MaxClient(
            phone=phone,
            work_dir=account_path,
            reconnect=False,  # Отключаем авто-переподключение при подключении
            headers=headers,
        )

        # Запускаем клиент (появится QR-код в консоли)
        console.print("\n[bold cyan]⏳ Запуск клиента...[/]")

        # Создаём задачу для запуска клиента в фоне
        async def run_client():
            try:
                await client.start()
            except Exception as e:
                # Игнорируем ошибку авторизации (ожидаема при подключении)
                if "FAIL_LOGIN_TOKEN" not in str(e):
                    raise

        client_task = asyncio.create_task(run_client())

        # Ждём пока клиент подключится (QR-код отобразится в консоли)
        await asyncio.sleep(2)

        # Проверяем, подключился ли клиент (получили данные профиля)
        if client.me:
            # Сохраняем аккаунт в базу данных
            try:
                MaxAccount.create(
                    phone=phone,
                    name=f"{client.me.first_name} {client.me.last_name}" if hasattr(client.me, 'first_name') else phone,
                    user_id=str(client.me.id) if client.me.id else None,
                    account_path=account_path,
                    is_active="Y",
                )
                console.print(f"\n[green]✅ Аккаунт {phone} успешно подключён![/]")
                console.print(f"[dim]ID пользователя: {client.me.id}[/]")
            except Exception as e:
                if "UNIQUE constraint" in str(e):
                    console.print(f"[yellow]⚠️ Аккаунт {phone} уже подключён[/]")
                else:
                    console.print(f"[red]❌ Ошибка сохранения: {e}[/]")
                    client_task.cancel()
                    return False
        else:
            # Если клиент ещё не получил данные профиля — ждём сканирования QR
            console.print("\n[yellow]⏳ Ожидание сканирования QR-кода...[/]")
            # Даём больше времени на сканирование (до 60 секунд)
            for i in range(30):
                await asyncio.sleep(2)
                if client.me:
                    break

            if client.me:
                # После сканирования сохраняем аккаунт в базу
                try:
                    MaxAccount.create(
                        phone=phone,
                        name=f"{client.me.first_name} {client.me.last_name}" if hasattr(client.me,
                                                                                        'first_name') else phone,
                        user_id=str(client.me.id) if client.me.id else None,
                        account_path=account_path,
                        is_active="Y",
                    )
                    console.print(f"\n[green]✅ Аккаунт {phone} успешно подключён![/]")
                except Exception as e:
                    if "UNIQUE constraint" not in str(e):
                        console.print(f"[red]❌ Ошибка сохранения: {e}[/]")
            else:
                console.print("[red]❌ Время ожидания истекло[/]")

        # Отменяем задачу клиента (подключение завершено)
        client_task.cancel()
        try:
            await client_task
        except asyncio.CancelledError:
            pass

        return client.me is not None

    except Exception as e:
        console.print(f"[red]❌ Ошибка подключения: {e}[/]")
        logger.exception("Ошибка при подключении аккаунта")
        return False


def show_accounts_list():
    """
    Показывает список подключённых аккаунтов в виде таблицы.
    
    Отображает: телефон, имя, ID, статус активности, дату подключения.
    """
    accounts = list(MaxAccount.select().order_by(MaxAccount.connected_at.desc()))

    if not accounts:
        # Если аккаунтов нет — показываем информационное сообщение
        console.print(Panel(
            "[yellow]📭 Пока нет подключённых аккаунтов[/]\n\n"
            "Используйте пункт [cyan][4][/cyan] для подключения.",
            title="[bold cyan]Аккаунты[/]",
            border_style="yellow",
            padding=(1, 3),
        ))
        return

    # Создаём таблицу для отображения аккаунтов
    table = Table(title="📱 Подключённые аккаунты", box=box.ROUNDED, border_style="cyan")
    table.add_column("Телефон", style="cyan", justify="center")
    table.add_column("Имя", style="green", justify="center")
    table.add_column("ID", style="dim", justify="center")
    table.add_column("Статус", style="yellow", justify="center")
    table.add_column("Подключён", style="dim", justify="center")

    for acc in accounts:
        # Статус: галочка если активен, крестик если нет
        status = "[green]✓[/green]" if acc.is_active == "Y" else "[red]✗[/red]"
        table.add_row(
            acc.phone,
            acc.name or "—",
            str(acc.user_id) if acc.user_id else "—",
            status,
            acc.connected_at.strftime("%d.%m.%Y %H:%M"),
        )

    console.print()
    console.print(table)
    console.print()


# ─── Заголовки Excel ──────────────────────────────────────────────────────────
# Словарь для перевода ключей данных на русский язык для Excel
HEADERS_RU = {
    'account_status': 'Статус аккаунта',
    'base_raw_url': 'Фото (raw URL)',
    'base_url': 'Фото (URL)',
    'description': 'Описание',
    'error': 'Ошибка',
    'gender': 'Пол',
    'id': 'ID пользователя',
    'link': 'Ссылка на профиль',
    'menu_button': 'Кнопка меню',
    'names': 'Имя',
    'options': 'Платформы',
    'photo_id': 'ID фото',
    'searched_phone': 'Искомый телефон',
    'update_time': 'Последняя активность',
    'web_app': 'Веб-приложение',
}


# ─── Утилиты ──────────────────────────────────────────────────────────────────

def extract_user_data(result) -> dict:
    """
    Извлекает данные пользователя из результата поиска MaxClient.
    
    Преобразует атрибуты объекта result в словарь, обрабатывая:
    - Временные метки (конвертирует из timestamp в дату)
    - Списки объектов (преобразует в строки)
    - Обычные значения
    
    :param result: Объект результата поиска.
    :return: Словарь с данными пользователя.
    """
    user_data = {}
    if not result or isinstance(result, (str, int, bool)):
        return user_data
    for attr in dir(result):
        if not attr.startswith('_'):  # Пропускаем приватные атрибуты
            try:
                val = getattr(result, attr)
                if not callable(val):  # Пропускаем методы
                    if attr == 'update_time' and isinstance(val, (int, float)):
                        # Конвертируем timestamp в читаемую дату
                        dt = datetime.fromtimestamp(val / 1000)
                        user_data[attr] = dt.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(val, list) and val and hasattr(val[0], '__dict__'):
                        # Списки объектов преобразуем в строки
                        user_data[attr] = [str(item) for item in val]
                    else:
                        user_data[attr] = val
            except Exception as e:
                logger.exception(e)
    return user_data


def save_to_excel(users_data: list[dict], filename: str = EXCEL_FILE) -> str:
    """
    Сохраняет данные пользователей в Excel файл.
    
    Если файл существует — добавляет новые данные, сохраняя существующие заголовки.
    Автоматически добавляет новые столбцы при их появлении.
    
    :param users_data: Список словарей с данными пользователей.
    :param filename: Путь к Excel файлу.
    :return: Путь к сохранённому файлу.
    """
    # Создаём директорию для файла (если не существует)
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    if os.path.exists(filename):
        # Файл существует — загружаем его
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        # Получаем существующие заголовки
        headers = [cell.value for cell in ws[1] if cell.value]
        ru_to_key = {v: k for k, v in HEADERS_RU.items()}  # Обратный словарь
        header_keys = [ru_to_key.get(h, h) for h in headers]
    else:
        # Создаём новый файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи"
        headers = []
        header_keys = []

    # Собираем все уникальные ключи из новых данных и существующих заголовков
    all_keys = set(header_keys)
    for user in users_data:
        all_keys.update(user.keys())

    new_keys = sorted(all_keys)
    # Если ключи изменились — обновляем заголовки
    if new_keys != header_keys:
        for col, key in enumerate(new_keys, 1):
            ws.cell(row=1, column=col, value=HEADERS_RU.get(key, key))
        header_keys = new_keys

    # Определяем первую пустую строку для добавления данных
    start_row = ws.max_row + 1 if ws.max_row > 1 else 2

    # Добавляем данные пользователей
    for row_idx, user in enumerate(users_data, start_row):
        for col_idx, key in enumerate(header_keys, 1):
            value = user.get(key, "")
            if isinstance(value, list):
                # Преобразуем списки в строку с разделителем
                value = " | ".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(filename)
    return filename


def load_numbers_to_db() -> int:
    """
    Загружает номера телефонов из файла в базу данных очереди.
    
    Пропускает дубликаты (использует get_or_create).
    
    :return: Общее количество номеров в очереди после загрузки.
    """
    numbers = read_file()  # Чтение номеров из файла (NUMBERS_FILE)
    added = 0
    for phone in numbers:
        try:
            # Добавляем номер, если он ещё не существует
            PhoneQueue.get_or_create(phone=str(phone).strip())
            added += 1
        except Exception as e:
            logger.exception(e)
    return PhoneQueue.select().count()


def get_queue_count() -> int:
    """
    Получает текущее количество номеров в очереди.
    
    :return: Количество записей в таблице PhoneQueue.
    """
    return PhoneQueue.select().count()


def remove_from_queue(phone: str):
    """
    Удаляет номер телефона из очереди после успешной обработки.
    
    :param phone: Номер телефона для удаления.
    """
    PhoneQueue.delete().where(PhoneQueue.phone == phone).execute()


def get_next_phones(batch: int = 1) -> list[str]:
    """
    Получает следующий пакет номеров из очереди.
    
    Номера возвращаются в порядке добавления (ASC по added_at).
    
    :param batch: Количество номеров для получения.
    :return: Список номеров телефонов.
    """
    rows = PhoneQueue.select().order_by(PhoneQueue.added_at).limit(batch)
    return [r.phone for r in rows]


# ─── UI ───────────────────────────────────────────────────────────────────────

def print_header():
    """
    Очищает консоль и выводит заголовок программы.
    
    Отображает название, описание и версию приложения.
    """
    console.clear()
    title = Text("📱  MAX / OneMe Phone Parser", style="bold cyan")
    subtitle = Text("Парсинг пользователей по номерам телефона", style="dim")
    version = Text(f"Версия: {APP_VERSION} от {APP_DATE}", style="yellow")
    console.print(Panel(
        Align.center(Text.assemble(title, "\n", subtitle, "\n", version)),
        border_style="cyan",
        padding=(1, 4),
    ))


def print_stats():
    """
    Выводит краткую статистику: количество номеров в очереди и файлы результатов.
    """
    total = get_queue_count()
    table = Table(box=box.ROUNDED, border_style="blue", show_header=False, padding=(0, 2))
    table.add_column(style="dim")
    table.add_column(style="bold green")
    table.add_row("📋 Осталось в очереди:", str(total))
    table.add_row("💾 Файл результатов:", EXCEL_FILE)
    table.add_row("⏱️ Задержка между запросами:", f"{SLEEP_TIME} сек")
    console.print(table)


def print_menu() -> str:
    """
    Выводит главное меню программы и запрашивает выбор пользователя.
    
    :return: Выбранный пункт меню (0-4).
    """
    console.print()
    console.print(Panel(
        "[bold white]Выберите действие:[/]\n\n"
        "  [cyan][1][/cyan] ▶  Продолжить / начать перебор номеров\n"
        "  [cyan][2][/cyan] 🔄 Обновить список номеров из файла\n"
        "  [cyan][3][/cyan] 📊 Показать статистику очереди\n"
        "  [cyan][4][/cyan] 🔌 Подключить аккаунт Max\n"
        "  [cyan][0][/cyan] ❌ Выйти",
        title="[bold cyan]Меню[/]",
        border_style="cyan",
        padding=(1, 3),
    ))
    choice = Prompt.ask(
        "[bold yellow]Ваш выбор[/]",
        choices=["0", "1", "2", "3", "4"],  # Доступные опции
        default="1",  # Выбор по умолчанию
    )
    return choice


# ─── Парсинг ──────────────────────────────────────────────────────────────────

async def parse_phones_with_rotation():
    """
    Перебор номеров с ротацией аккаунтов.
    
    Основная функция парсинга, которая:
    1. Получает активные аккаунты из базы
    2. Подключается к первому доступному аккаунту
    3. Перебирает номера из очереди по одному
    4. Ищет пользователей по номеру телефона
    5. Сохраняет результаты в Excel
    6. При блокировке аккаунта переключается на следующий
    7. Обрабатывает rate limit ограничения
    
    Если аккаунт получает блокировку — переключается на следующий.
    Использует client_connect() для подключения.
    """
    # Получаем активные аккаунты из базы данных
    accounts = get_active_accounts()

    # Проверяем наличие доступных аккаунтов
    if not accounts:
        console.print(Panel(
            "[bold red]❌ Нет доступных аккаунтов![/]\n\n"
            "Подключите аккаунты через пункт [cyan][4][/cyan] меню.",
            title="[bold yellow]Внимание[/]",
            border_style="red",
            padding=(1, 3),
        ))
        return

    # Получаем общее количество номеров в очереди
    total_start = get_queue_count()
    if total_start == 0:
        console.print(Panel(
            "[yellow]Очередь пуста.[/]\nЗагрузите номера через пункт [cyan][2][/cyan] меню.",
            border_style="yellow",
        ))
        return

    # Вывод сообщения о начале обработки
    console.print(Panel(
        f"[green]Начинаем обработку [bold]{total_start}[/bold] номеров...[/]\n"
        f"[dim]Аккаунтов доступно: {len(accounts)}[/]\n"
        f"[dim]Нажмите Ctrl+C для паузы[/]",
        border_style="green",
        padding=(1, 3),
    ))

    # Счётчики для статистики
    processed = 0  # Обработано номеров
    found = 0  # Найдено пользователей
    errors = 0  # Количество ошибок
    account_switches = 0  # Переключений между аккаунтами
    consecutive_errors = 0  # Счётчик последовательных ошибок
    current_account_index = 0  # Индекс текущего аккаунта в списке
    current_account = None  # Текущий объект аккаунта

    try:
        # Подключаем первый аккаунт через client_connect
        current_account = accounts[current_account_index]
        console.print(f"\n[bold cyan]🔌 Подключение к аккаунту: {current_account.phone}[/]")

        # Подключение к аккаунту (с кэшированием)
        current_client = await client_connect(
            phone=current_account.phone,
            work_dir=current_account.account_path
        )

        # Отмечаем аккаунт как работающий
        mark_account_working(current_account.phone)
        log_account_action(current_account.phone, "start", "Начал работу")

        # Ждём пока загрузится профиль пользователя (до 2 секунд)
        for _ in range(20):
            await asyncio.sleep(0.1)
            if current_client.me and current_client.me.id:
                break

        # Отображаем информацию о текущем аккаунте
        account_info = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
        account_info.add_column(style="dim")
        account_info.add_column(style="bold cyan")
        me = current_client.me
        account_info.add_row("Текущий аккаунт:", current_account.phone)
        account_info.add_row("ID пользователя:", str(me.id) if me and me.id else "загрузка...")

        # Получаем имя пользователя из различных полей профиля
        if me:
            name_parts = []
            if hasattr(me, 'first_name') and me.first_name:
                name_parts.append(me.first_name)
            if hasattr(me, 'last_name') and me.last_name:
                name_parts.append(me.last_name)
            if hasattr(me, 'names') and me.names:
                # Если есть список имён — берём первое
                if isinstance(me.names, list) and len(me.names) > 0:
                    name_parts.append(str(me.names[0]))

            name = " ".join(name_parts) if name_parts else "—"
        else:
            name = "—"

        account_info.add_row("Имя:", name)
        account_info.add_row("Ошибок аккаунта:", current_account.errors_count)
        console.print(account_info)
        console.print()

        # Создаём прогресс-бар для отображения хода обработки
        with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                      BarColumn(), TaskProgressColumn(), TextColumn("[dim]{task.fields[status]}[/]"),
                      console=console, transient=False) as progress:

            task = progress.add_task(
                "[cyan]Обработка...",
                total=total_start,
                status="",
            )

            # Основной цикл обработки номеров
            while True:
                # Получаем следующий номер из очереди
                phones = get_next_phones(1)
                if not phones:
                    break  # Очередь пуста

                phone = phones[0]
                progress.update(task, description=f"[cyan]📞 {phone}", status="запрос...")

                # Задержка между запросами (защита от rate limit)
                await asyncio.sleep(SLEEP_TIME)

                try:
                    # Поиск пользователя по номеру с автоматическим переподключением
                    result = await safe_search(phone=phone, client=current_client)

                    # Если результат None — ошибка соединения
                    if result is None:
                        logger.error(f"Не удалось получить данные для {phone} после нескольких попыток")
                        status_text = "❌ нет соединения"
                        consecutive_errors += 1

                        # Если много ошибок подряд — переключаем аккаунт
                        if consecutive_errors >= 3:
                            console.print(f"\n[yellow]⚠️ Много ошибок подряд, переключаем аккаунт...[/]")
                            is_blocked = mark_account_error(current_account.phone)

                            # Переключаем на следующий аккаунт
                            if current_account_index < len(accounts) - 1:
                                # Отключаем текущего клиента
                                await client_disconnect(current_account.phone)

                                current_account_index += 1
                                current_account = accounts[current_account_index]

                                console.print(f"\n[bold cyan]🔄 Переключение на аккаунт: {current_account.phone}[/]")

                                # Подключаем новый аккаунт через client_connect
                                current_client = await client_connect(
                                    phone=current_account.phone,
                                    work_dir=current_account.account_path
                                )

                                mark_account_working(current_account.phone)
                                account_switches += 1
                                consecutive_errors = 0
                                status_text = f"🔄 аккаунт переключён"
                            else:
                                console.print("[yellow]⚠️ Больше нет доступных аккаунтов[/]")
                                consecutive_errors = 0
                        continue

                    consecutive_errors = 0  # Сбрасываем счётчик ошибок

                    # Если пользователь найден — извлекаем данные и сохраняем
                    if result and not isinstance(result, (str, int, bool)):
                        user_data = extract_user_data(result)
                        user_data['searched_phone'] = phone  # Добавляем искомый номер
                        save_to_excel([user_data], EXCEL_FILE)
                        found += 1
                        status_text = f"✅ найден: {user_data.get('names', ['?'])[0] if user_data.get('names') else '?'}"
                        logger.info(f"Найден пользователь {phone}: {user_data}")
                    else:
                        # Пользователь не найден в Max
                        status_text = "⚪ не найден"
                        logger.info(f"Пользователь не найден: {phone}")

                except Exception as e:
                    err_str = str(e)

                    # Проверяем на rate limit (слишком много запросов)
                    if 'too-many' in err_str.lower() or 'ratelimit' in err_str.lower():
                        progress.update(task, status=f"⏳ rate limit, ждём {SLEEP_ON_RATELIMIT}с...")
                        logger.warning(f"Rate limit на {phone}, ждём {SLEEP_ON_RATELIMIT}с")
                        console.print(f"[yellow]⏳ Rate limit, ожидание {SLEEP_ON_RATELIMIT}с...[/]")
                        await asyncio.sleep(SLEEP_ON_RATELIMIT)
                        progress.update(task, status="продолжение...")
                        continue

                    # Проверяем на ошибку авторизации/блокировку аккаунта
                    if 'FAIL_LOGIN_TOKEN' in err_str or 'авторизируйтесь' in err_str.lower():
                        console.print(f"\n[red]❌ Аккаунт {current_account.phone} заблокирован![/]")
                        mark_account_blocked(current_account.phone, "FAIL_LOGIN_TOKEN")

                        # Отключаем заблокированного клиента
                        await client_disconnect(current_account.phone)

                        # Переключаем на следующий аккаунт
                        if current_account_index < len(accounts) - 1:
                            current_account_index += 1
                            current_account = accounts[current_account_index]

                            console.print(f"\n[bold cyan]🔄 Переключение на аккаунт: {current_account.phone}[/]")

                            # Подключаем новый аккаунт через client_connect
                            current_client = await client_connect(
                                phone=current_account.phone,
                                work_dir=current_account.account_path
                            )

                            mark_account_working(current_account.phone)
                            account_switches += 1
                            status_text = f"🔄 аккаунт переключён (блокировка)"
                            continue
                        else:
                            console.print("[red]❌ Нет доступных аккаунтов для продолжения[/]")
                            break

                    # Проверяем на "номер не найден" — это не ошибка аккаунта
                    if 'not found' in err_str.lower() or 'не найдено' in err_str.lower():
                        status_text = "⚪ не найден в Max"
                        logger.info(f"Номер {phone} не найден в базе Max")
                        # Не считаем это ошибкой, просто идём дальше

                    # Другая ошибка — сохраняем в Excel и регистрируем как ошибку аккаунта
                    else:
                        error_data = {'searched_phone': phone, 'error': f"{type(e).__name__}: {e}"}
                        save_to_excel([error_data], EXCEL_FILE)
                        errors += 1
                        status_text = f"❌ ошибка"
                        logger.error(f"Ошибка для {phone}: {e}")

                        # Регистрируем ошибку аккаунта
                        is_blocked = mark_account_error(current_account.phone)
                        if is_blocked and current_account_index < len(accounts) - 1:
                            console.print(f"\n[yellow]⚠️ Аккаунт заблокирован после ошибок[/]")

                            # Отключаем текущего клиента
                            await client_disconnect(current_account.phone)

                            current_account_index += 1
                            current_account = accounts[current_account_index]

                            # Подключаем новый аккаунт через client_connect
                            current_client = await client_connect(
                                phone=current_account.phone,
                                work_dir=current_account.account_path
                            )

                            mark_account_working(current_account.phone)
                            account_switches += 1

                # Удаляем из очереди только после успешной обработки
                remove_from_queue(phone)
                processed += 1
                progress.update(task, advance=1, status=status_text)

    except KeyboardInterrupt:
        # Обработка прерывания пользователем (Ctrl+C)
        console.print("\n[yellow]⏸ Остановлено пользователем[/]")
    except Exception as e:
        # Обработка критических ошибок
        console.print(f"\n[red]❌ Критическая ошибка: {e}[/]")
        logger.exception("Критическая ошибка в parse_phones_with_rotation")
    finally:
        # Завершение работы: отключаем всех клиентов и освобождаем аккаунт
        await disconnect_all_clients()
        if current_account:
            mark_account_idle(current_account.phone)

    # Вывод итоговой статистики
    console.print()
    summary_table = Table(box=box.ROUNDED, border_style="green", title="📊 Результаты")
    summary_table.add_column("Параметр", style="dim")
    summary_table.add_column("Значение", style="bold")
    summary_table.add_row("Обработано", str(processed))
    summary_table.add_row("Найдено", str(found))
    summary_table.add_row("Ошибок", str(errors))
    summary_table.add_row("Переключений аккаунтов", str(account_switches))
    summary_table.add_row("Результаты", EXCEL_FILE)

    console.print(summary_table)


# Для обратной совместимости
async def parse_phones(client=None):
    """Обёртка для старой версии (если вызывается с одним клиентом)."""
    await parse_phones_with_rotation()


async def safe_search(phone: str, client, retries: int = 3):
    """
    Поиск с автоматическим переподключением.

    Пытается выполнить поиск пользователя по номеру телефона.
    При ошибке WebSocket/not connected ждёт переподключения и повторяет попытку.

    :param phone: Номер телефона для поиска.
    :param client: Экземпляр MaxClient.
    :param retries: Количество попыток при ошибке подключения.
    :return: Результат поиска или None если все попытки не удались.
    """
    for attempt in range(retries):
        try:
            result = await client.search_by_phone(phone=phone)
            return result
        except Exception as e:
            err = str(e).lower()
            if 'not connected' in err or 'websocket' in err:
                # Ошибка соединения — ждём переподключения
                logger.warning(f"WebSocket отключён, ждём переподключения... (попытка {attempt + 1}/{retries})")
                await asyncio.sleep(10)  # даём pymax время переподключиться
                continue
            raise  # остальные ошибки — пробрасываем
    return None


# ─── Точка входа ──────────────────────────────────────────────────────────────


async def main():
    """
    Основное меню программы.
    
    Запускает бесконечный цикл меню с обработкой выбора пользователя:
    0 - Выход из программы
    1 - Запуск парсинга номеров
    2 - Загрузка номеров из файла
    3 - Показать статистику очереди
    4 - Подключить новый аккаунт Max
    """

    print_header()

    while True:
        print_stats()
        choice = print_menu()

        if choice == "0":  # Выход из программы
            console.print("\n[dim]До свидания![/]")
            sys.exit(0)

        elif choice == "1":  # Запуск парсинга номеров с переключением аккаунтов
            try:
                await parse_phones_with_rotation()
            except KeyboardInterrupt:
                console.print("\n[yellow]⏸  Пауза. Прогресс сохранён в БД.[/]")

        elif choice == "2":  # Загрузка номеров из файла
            with console.status("[cyan]Загружаем номера из файла...[/]"):
                count = load_numbers_to_db()
            console.print(Panel(
                f"[green]✅ Готово![/] В очереди теперь [bold]{count}[/bold] номеров.",
                border_style="green",
            ))

        elif choice == "3":  # Показать статистику очереди
            print_header()
            # Детальная статистика
            total = get_queue_count()
            table = Table(title="Статистика очереди", box=box.ROUNDED, border_style="cyan")
            table.add_column("Параметр", style="dim")
            table.add_column("Значение", style="bold")
            table.add_row("Номеров в очереди", str(total))
            table.add_row("Файл номеров", NUMBERS_FILE)
            table.add_row("База данных", DB_PATH)
            table.add_row("Результаты Excel", EXCEL_FILE)
            table.add_row("Задержка запросов", f"{SLEEP_TIME} сек")
            table.add_row("Задержка при rate limit", f"{SLEEP_ON_RATELIMIT} сек")
            console.print(table)

        elif choice == "4":  # Подключить новый аккаунт
            print_header()

            show_accounts_list()  # Показываем список подключённых аккаунтов

            connect = Prompt.ask(  # Предлагаем подключить новый аккаунт
                "[bold yellow]Подключить новый аккаунт?[/]",
                choices=["y", "n"],
                default="y",
            )

            if connect.lower() == "y":
                await connect_account_by_qr()

        console.print()
        input("  [Enter] для возврата в меню...")
        print_header()


if __name__ == "__main__":
    asyncio.run(main())
